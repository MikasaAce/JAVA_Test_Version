import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd
import traceback
from app.api.config import config as data_class

# 获取py 文件所在目录
current_path = os.path.dirname(__file__)

# 把这个目录设置成工作目录
os.chdir(current_path)

csv_path = data_class.pdf_save_path



def draw_sheet_1(ws, info_labels, data):
    # ================== 样式配置 ==================
    # 颜色定义（修正命名）
    background_light_gray = PatternFill(start_color='D5DAE6', fill_type='solid')
    background_dark_gray = PatternFill(start_color='95B3D7', fill_type='solid')
    background_yellow = PatternFill(start_color='FFFF00', fill_type='solid')

    # 对齐方式
    cell_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True) # 启用文本换行
    cell_alignment_center = Alignment(horizontal='center', vertical='center')

    # 边框样式
    thin_border = Border(
        left=Side(style='thin', color='808080'),
        right=Side(style='thin', color='808080'),
        top=Side(style='thin', color='808080'),
        bottom=Side(style='thin', color='808080')
    )

    # ================== 基础信息表格 ==================
    # 设置列宽
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 24

    # 填充基础信息
    for row, (label, value) in enumerate(info_labels, start=1):
        ws[f'A{row}'] = label
        ws[f'A{row}'].fill = background_dark_gray
        ws[f'A{row}'].border = thin_border
        ws[f'A{row}'].alignment = cell_alignment_left

        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = background_light_gray
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].alignment = cell_alignment_left

    start_line = 8
    if len(info_labels) == 9:
        start_line = 10

    # ================== 合并单元格 ==================
    ws.merge_cells(f'A{start_line}:C{start_line}')
    ws[f'A{start_line}'] = '漏洞检测结果详情统计'
    ws[f'A{start_line}'].fill = background_yellow
    ws[f'A{start_line}'].border = thin_border
    ws[f'A{start_line}'].alignment = cell_alignment_center  # 居中对齐

    # 表头
    headers = ['序号', '漏洞名称', '统计']
    for col, header in zip(['A', 'B', 'C'], headers):
        cell = ws[f'{col}{start_line+1}']
        cell.value = header
        cell.fill = background_yellow
        cell.border = thin_border
        cell.alignment = cell_alignment_left

    # ================== 填充数据 ==================
    for row, (data1, data2, data3) in enumerate(data, start=start_line+2):
        # A列
        ws[f'A{row}'].value = data1
        ws[f'A{row}'].fill = background_dark_gray
        ws[f'A{row}'].border = thin_border
        ws[f'A{row}'].alignment = cell_alignment_left

        # B列
        ws[f'B{row}'].value = data2
        ws[f'B{row}'].fill = background_light_gray  # 保持与基础信息区一致
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].alignment = cell_alignment_left

        # C列
        ws[f'C{row}'].value = data3
        ws[f'C{row}'].fill = background_light_gray
        ws[f'C{row}'].border = thin_border
        ws[f'C{row}'].alignment = cell_alignment_left


def draw_sheet_2(ws, data):
    # ================== 样式配置 ==================
    # 颜色定义
    fill = {
        'yellow': PatternFill(start_color='FFFF00', fill_type='solid'),
        'dark_gray': PatternFill(start_color='95B3D7', fill_type='solid'),
        'light_gray': PatternFill(start_color='D5DAE6', fill_type='solid')
    }

    # 对齐方式
    alignment = {
        'left': Alignment(horizontal='left', vertical='center'),
        'center': Alignment(horizontal='center', vertical='center')
    }

    # 边框样式
    thin_border = Border(
        left=Side(style='thin', color='808080'),
        right=Side(style='thin', color='808080'),
        top=Side(style='thin', color='808080'),
        bottom=Side(style='thin', color='808080')
    )

    # ================== 表格结构配置 ==================
    # 列宽配置（列字母: 宽度）
    column_widths = {
        'A': 12, 'B': 12, 'C': 36, 'D': 12, 'E': 36,
        'F': 36, 'G': 36, 'H': 36, 'I': 36, 'J': 12, 'K': 36
    }

    # 表头结构配置（合并范围，标题，填充颜色）
    header_merge_ranges = [
        (f'B{{row}}:F{{row}}', '漏洞概述', fill['yellow']),
        (f'G{{row}}:K{{row}}', '漏洞详情', fill['yellow'])
    ]

    # 二级表头（列字母，标题）
    sub_headers = [
        ('B', '漏洞编号'), ('C', '漏洞名称'), ('D', '风险等级'), ('E', '漏洞描述'),
        ('F', '修复建议'), ('G', '漏洞文件'), ('H', '漏洞爆发点'), ('I', '爆发点函数'),
        ('J', '缺陷源'), ('K', '漏洞源代码')
    ]

    # ================== 初始化表格样式 ==================
    # 设置列宽
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    row_current = 1
    for vul_idx, (vul_id, name, level, description, suggestion, details) in enumerate(data, start=1):
        # ================== 主表头 ==================
        # 设置左侧编号单元格
        ws[f'A{row_current}'] = f'vul{vul_idx}'
        ws[f'A{row_current}'].fill = fill['yellow']
        ws[f'A{row_current}'].border = thin_border
        ws[f'A{row_current}'].alignment = alignment['center']

        # 处理合并表头
        for range_pattern, title, color in header_merge_ranges:
            merge_range = range_pattern.format(row=row_current)
            ws.merge_cells(merge_range)

            # 设置合并区域样式
            for row in ws[merge_range]:
                for cell in row:
                    cell.fill = color
                    cell.border = thin_border
                    cell.alignment = alignment['center']

            # 写入标题到左上角单元格
            ws[merge_range.split(':')[0]].value = title

        # ================== 二级表头 ==================
        row_current += 1
        for col, title in sub_headers:
            cell = ws[f'{col}{row_current}']
            cell.value = title
            cell.fill = fill['dark_gray']
            cell.border = thin_border
            cell.alignment = alignment['left']

        # ================== 基础数据行 ==================
        row_current += 1
        base_data = [vul_id, name, level, description, suggestion]
        for col, value in zip(['B', 'C', 'D', 'E', 'F'], base_data):
            cell = ws[f'{col}{row_current}']
            cell.value = value
            cell.border = thin_border
            cell.alignment = alignment['left']

        # ================== 详情数据行 ==================
        for file, point, func, source, code in details:
            file_extension = os.path.splitext(file)[1]
            if file_extension not in ['.java', '.class', '.xml', '.js', '.m', '.h', '.go', '.py', '.pyc',
                                      '.c', '.cpp', '.php', '.rb', '.sql']:
                continue
            detail_data = [file, point, func, source, code]
            for col, value in zip(['G', 'H', 'I', 'J', 'K'], detail_data):
                cell = ws[f'{col}{row_current}']
                cell.value = value
                cell.border = thin_border
                cell.alignment = alignment['left']
            row_current += 1

        # 更新到下一个漏洞起始行（留空一行）
        row_current += 1


def export_excel1(itemName, excel_Time, zipName, vulFileNumber, language, detect_type, startTime, lastTime,
                   vul_number, header_five, vuls, file_location, final_results, number, version, risk_level_dict,git_info):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = itemName

        # ================== 第一张表 ==================
        # 填充基础数据
        total = sum(map(lambda x: x[1], number))
        info_labels = [
            ('项目名称', itemName),
            ('检测文件名称', zipName),
            ('有漏洞文件数量', total),
            ('开发语言', language),
            ('扫描类型', detect_type),
            ('检测开始时间', startTime),
            ('检测耗时', lastTime)
        ]
        if git_info[0]:
            info_labels.extend([
                ('git仓库地址', f'{git_info[0] or ""}'),
                ('分支', f'{git_info[1] or "master"}'),
            ])

        summary_data = []
        for i in range(0, len(vul_number)):
            summary_data.append(
                ('{}'.format(int(vul_number[i][0])), '{}'.format(vul_number[i][1]), '{}'.format(vul_number[i][2])))

        draw_sheet_1(ws, info_labels, summary_data)

        # ================== 第二张表 ==================
        #region 获取数据
        all_data = []
        if version is not None:
            if 'Developer Workbook' in version:
                vuls = vuls.split(',')
                num = 0
                for i, vul in enumerate(vuls):
                    # 拿到这里的这些数据，看看具体都是什么东西，然后才好针对性的写代码
                    vul = vul.strip()
                    level = risk_level_dict[vul]
                    Description = None
                    suggestion = None
                    details = []
                    for i2 in range(0, number[i][1]):
                        id = num + i2
                        if vul.lower().replace(' ', '') == file_location[id][1].replace(" ", ""):
                            detail = (
                                file_location[id][0],
                                file_location[id][5],
                                file_location[id][6],
                                file_location[id][7],
                                file_location[id][3],
                            )
                            details.append(detail)
                    num += int(number[i][1])

                    data = (
                        number[i][1],
                        vul,
                        level,
                        Description,
                        suggestion,
                        details
                    )

                    all_data.append(data)
            elif 'OWASP' in version:
                if '2010' in version:
                    table = pd.read_csv(os.path.join(csv_path, "csv/OWASP_2010.csv"), encoding='gbk')
                elif '2013' in version:
                    table = pd.read_csv(os.path.join(csv_path, "csv/OWASP_2013.csv"), encoding='gbk')
                elif '2017' in version:
                    table = pd.read_csv(os.path.join(csv_path, "csv/OWASP_2017.csv"), encoding='gbk')
                vul_names = table['漏洞名称'].values.tolist()
                levels = table['漏洞等级'].values.tolist()
                Descriptions = table['漏洞描述'].values.tolist()
                suggestions = table['修复方案'].values.tolist()
                vuls = vuls.split(',')
                num = 0
                for i, vul in enumerate(vuls):
                    id_front = 0
                    for kkk, _ in enumerate(vul_names):
                        if _ == vul:
                            id_front = kkk
                    vul_name = vul_names[id_front]
                    level = levels[id_front]
                    Description = Descriptions[id_front]
                    suggestion = suggestions[id_front]
                    details = []
                    for i2 in range(0, number[i][1]):
                        id = num + i2
                        if vul.lower().replace(' ', '') == file_location[id][1].replace(" ", ""):
                            detail = (
                                file_location[id][0],
                                file_location[id][5],
                                file_location[id][6],
                                file_location[id][7],
                                file_location[id][3],
                            )
                            details.append(detail)
                    num += int(number[i][1])

                    data = (
                        number[i][1],
                        vul_name,
                        level,
                        Description,
                        suggestion,
                        details
                    )

                    all_data.append(data)
            elif 'CWE' in version:
                table = pd.read_csv(os.path.join(csv_path, "csv/漏洞知识库.csv"), encoding='gbk')
                ids = table['漏洞ID'].values.tolist()
                cweNames = table['漏洞名称'].values.tolist()
                levels = table['漏洞等级'].values.tolist()
                Descriptions = table['漏洞描述'].values.tolist()
                suggestions = table['修复方案'].values.tolist()
                vuls = vuls.split(',')
                num = 0
                for i, vul in enumerate(vuls):
                    id_front = 0
                    for kkk, _ in enumerate(ids):
                        if _ == vul.lower().strip():
                            id_front = kkk
                    cweName = cweNames[id_front]
                    level = levels[id_front]
                    Description = Descriptions[id_front]
                    suggestion = suggestions[id_front]
                    details = []
                    for i2 in range(0, number[i][1]):
                        id = num + i2
                        if vul.lower().replace(' ', '') == file_location[id][1].replace(" ", ""):
                            detail = (
                                file_location[id][0],
                                file_location[id][5],
                                file_location[id][6],
                                file_location[id][7],
                                file_location[id][3],
                            )
                            details.append(detail)
                    num += int(number[i][1])

                    data = (
                        number[i][1],
                        cweName,
                        level,
                        Description,
                        suggestion,
                        details
                    )

                    all_data.append(data)

        else:
            for index, (vul_name, level, description, repairPlan, file_list) in enumerate(final_results, start=1):
                details = []
                for file_index, (
                fileName, vulType, location, source_code, repair_code, Sink, Enclosing_Method, Source) in enumerate(
                        file_list, start=1):
                    detail = (
                        fileName,
                        Sink,
                        Enclosing_Method,
                        Source,
                        source_code
                    )
                    details.append(detail)

                data = (
                    len(file_list),
                    vul_name,
                    level,
                    description,
                    repairPlan,
                    details
                )
                all_data.append(data)
            # table = pd.read_csv(os.path.join(csv_path, "csv/漏洞知识库_subVul.csv"), encoding='gbk')
            # vul_names = table['漏洞名称'].values.tolist()
            # levels = table['漏洞等级'].values.tolist()
            # Descriptions = table['漏洞描述'].values.tolist()
            # suggestions = table['修复方案'].values.tolist()
            # vuls = vuls.split(',')
            # num = 0
            # for i, vul in enumerate(vuls):
            #     id_front = 0
            #     # 找到id
            #     for kkk, _ in enumerate(vul_names):
            #         if _ == vul.strip():
            #             id_front = kkk
            #             break
            #     vul_name = vul_names[id_front]
            #     level = levels[id_front]
            #     Description = Descriptions[id_front]
            #     suggestion = suggestions[id_front]
            #     details = []
            #     for i2 in range(0, number[i][1]):
            #         id = num + i2
            #         if vul.lower().replace(' ', '') == file_location[id][1].replace(" ", ""):
            #             detail = (
            #                 file_location[id][0],
            #                 file_location[id][5],
            #                 file_location[id][6],
            #                 file_location[id][7],
            #                 file_location[id][3],
            #             )
            #             details.append(detail)
            #     num += int(number[i][1])
            #
            #     data = (
            #         number[i][1],
            #         vul_name,
            #         level,
            #         Description,
            #         suggestion,
            #         details
            #     )
            #
            #     all_data.append(data)


        # endregion

        ws1 = wb.create_sheet('details')
        draw_sheet_2(ws1, all_data)


        # ================== 保存文件 ==================
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{itemName}_检测报告_{timestamp}.xlsx"
        wb.save(os.path.join(csv_path, filename))
        print(f"报告已生成：{filename} in {csv_path}/{filename}_{timestamp}.xlsx")
        return filename
    except Exception as e:
        traceback.print_exc()
        print(f'导出excel文件时出错: {e}')


